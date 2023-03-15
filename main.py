import pandas as pd
from stocks import StockHandler, Stock, Screener
from alive_progress import alive_bar
import openpyxl
from openpyxl.styles import PatternFill

str_ax = '.ax'
url = "https://www.asx.com.au/asx/research/ASXListedCompanies.csv"
handler = None
rsi_cols = ['Price', 'RSI', 'RSI_Signal', 'MACD', 'signal', 'hist', 'MACD_Signal', 'Active']
sheet_name = 'Stock stats'

def get_stock_list():
    handler = StockHandler()
    handler.get_asx_stock_list(url)

    # FOrmat the output collection for processing.
    handler.format_collection()
    
def get_RSI_MACD_data():
    
    df_hist = pd.read_csv('data/stocks.csv', 
                          usecols=["asx_code", "company_name", "sector"])
    df = df_hist.reindex(columns=df_hist.columns.tolist() + rsi_cols)

    with alive_bar(df.shape[0],
                   ctrl_c=False, 
                   bar='smooth', 
                   spinner='classic', 
                   title='Loading...') as bar:
        
        for index, row in df.iterrows():

            ticker = row["asx_code"] + str_ax 
            
            try:
                screener = Screener(Stock(ticker))
                stock = screener.getStockStats()
        
                df.loc[index, ['Price']] = round(float(stock.price), 2)
                df.loc[index, ['RSI']] = round(float(stock.rsi), 2)
                df.loc[index, ['MACD']] = round(float(stock.macd), 3)
                df.loc[index, ['signal']] = round(float(stock.signal), 3)   
                df.loc[index, ['hist']] = round(float(stock.hist), 3)                           
                df.loc[index, ['RSI_Signal']] = stock.rsi_sign  
                df.loc[index, ['MACD_Signal']] = stock.macd_sign  
                df.loc[index, ['Active']] = 'Y'
                
            except Exception as exc:
                df.loc[index, ['Active']] = 'N'
                continue

            bar()

    df = df[(df.Active == 'Y')]
    print("Completed successfully..")  
    df.to_excel('data/rsi.xlsx', sheet_name=sheet_name, index=False) 

    # Fill excel signal indicator 
    fill_stock_signals()

def fill_stock_signals():

    # open excel
    fill_cell = PatternFill(patternType='solid', fgColor='ffff00')
    wb = openpyxl.load_workbook('data/rsi.xlsx')
    ws = wb[sheet_name] 

    column_names = []
    for cell in ws[1]:
        column_names.append(cell.value)
    rsi_signal_col_index = column_names.index('RSI_Signal') + 1
    macd_signal_col_index = column_names.index('MACD_Signal') + 1

    maxRow = ws.max_row
    maxCol = ws.max_column

    for rowNum in range(2, maxRow + 1):
        rsi_val = ws.cell(row=rowNum, column=rsi_signal_col_index).value
        macd_val = ws.cell(row=rowNum, column=macd_signal_col_index).value

        if rsi_val != None:
            ws.cell(row=rowNum, column=rsi_signal_col_index).value = ''
            ws.cell(row=rowNum, column=rsi_signal_col_index).fill = PatternFill(patternType='solid', fgColor=rsi_val)

        if macd_val != None:
            ws.cell(row=rowNum, column=macd_signal_col_index).value = ''
            ws.cell(row=rowNum, column=macd_signal_col_index).fill = PatternFill(patternType='solid', fgColor=macd_val)

    wb.save('data/rsi.xlsx')

def main():
    # # download current stock list
    # get_stock_list()

    # # Get RSI score for the active stocks.
    get_RSI_MACD_data()


if __name__ == '__main__':
    # print('pandas version: ', pd.__version__)
    # print('yfinance version: ', yf.__version__)
    main()


